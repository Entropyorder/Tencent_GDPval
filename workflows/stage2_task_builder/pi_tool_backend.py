#!/usr/bin/env python3
"""Deterministic backend for the Stage 2 Pi custom tools."""

import argparse
import csv
from contextlib import contextmanager
import fcntl
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import sys
import tempfile
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


WORKFLOW_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(WORKFLOW_DIR))

from validate_task_output import validate_task


GENERATED_EXTENSIONS = {
    "markdown": ".md",
    "text": ".txt",
    "csv": ".csv",
    "xlsx": ".xlsx",
}
ROLES = {"core", "supporting", "purposeful_noise", "generated"}
ATTACHMENT_PREFIX_PATTERN = re.compile(r"^\d{2}__")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(path)


def clean_filename(value: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError("filename must be a non-empty string")
    filename = value.strip()
    if Path(filename).name != filename or filename in {".", ".."}:
        raise ValueError("filename must not contain a directory path")
    if filename.startswith(".") or "\x00" in filename:
        raise ValueError("filename is unsafe")
    return filename


def final_attachment_filename(position: int, source_filename: str) -> str:
    base = ATTACHMENT_PREFIX_PATTERN.sub("", clean_filename(source_filename))
    return f"{position:02d}__{base}"


def task_context(value: str) -> tuple[Path, dict[str, Any]]:
    task_dir = Path(value).resolve()
    if not re.fullmatch(r"task_\d{3}", task_dir.name):
        raise ValueError("Pi tools must run from a task_NNN directory")
    manifest_path = task_dir / "candidate_manifest.json"
    if not manifest_path.is_file():
        raise FileNotFoundError(f"missing candidate manifest: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    candidates = manifest.get("candidates", [])
    if len(candidates) != 20:
        raise ValueError("candidate manifest must contain exactly 20 files")
    return task_dir, manifest


def candidate_by_rank(manifest: dict[str, Any], rank: int) -> dict[str, Any]:
    for candidate in manifest["candidates"]:
        if int(candidate["rank"]) == rank:
            return candidate
    raise ValueError(f"candidate rank not found: {rank}")


def candidate_by_id(
    manifest: dict[str, Any], document_id: str
) -> dict[str, Any]:
    for candidate in manifest["candidates"]:
        if candidate["document_id"] == document_id:
            return candidate
    raise ValueError(f"candidate document_id not found: {document_id}")


def local_path(task_dir: Path, relative: str) -> Path:
    candidate = task_dir / relative
    lexical = Path(os.path.abspath(candidate))
    if not lexical.is_relative_to(task_dir):
        raise ValueError(f"path leaves the current task workspace: {relative}")
    return candidate


@contextmanager
def mutation_lock(task_dir: Path):
    lock_path = task_dir / ".pi-tool.lock"
    with lock_path.open("a+", encoding="utf-8") as stream:
        fcntl.flock(stream.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(stream.fileno(), fcntl.LOCK_UN)


def inventory(task_dir: Path, manifest: dict[str, Any], _: dict[str, Any]):
    rows = [
        {
            "rank": int(item["rank"]),
            "document_id": item["document_id"],
            "filename": item["attachment_filename"],
            "document_type": item.get("document_type"),
            "subject": item.get("subject_name"),
            "business_topic": item.get("business_topic"),
            "summary": str(item.get("summary") or "")[:900],
            "extracted_path": item["extracted_path"],
        }
        for item in manifest["candidates"]
    ]
    payload = {
        "task_id": manifest["task_id"],
        "original_query": manifest.get("original_query"),
        "retrieval_query": manifest.get("retrieval_query"),
        "candidates": rows,
    }
    return {
        "message": json.dumps(payload, ensure_ascii=False, indent=2),
        "details": {"candidate_count": len(rows)},
    }


def read_candidate(
    task_dir: Path, manifest: dict[str, Any], params: dict[str, Any]
):
    rank = int(params["rank"])
    start_line = max(1, int(params.get("startLine", 1)))
    line_count = min(400, max(1, int(params.get("lineCount", 160))))
    candidate = candidate_by_rank(manifest, rank)
    path = local_path(task_dir, candidate["extracted_path"])
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    selected = lines[start_line - 1 : start_line - 1 + line_count]
    numbered = "\n".join(
        f"L{line_number}: {line}"
        for line_number, line in enumerate(selected, start=start_line)
    )
    return {
        "message": (
            f"rank={rank} document_id={candidate['document_id']} "
            f"lines={start_line}-{start_line + len(selected) - 1}\n{numbered}"
        ),
        "details": {
            "rank": rank,
            "total_lines": len(lines),
            "returned_lines": len(selected),
        },
    }


def search_evidence(
    task_dir: Path, manifest: dict[str, Any], params: dict[str, Any]
):
    terms = [
        str(term).strip()
        for term in params.get("terms", [])
        if str(term).strip()
    ]
    if not 1 <= len(terms) <= 12:
        raise ValueError("search_evidence requires 1-12 non-empty terms")
    max_hits = min(80, max(1, int(params.get("maxHits", 40))))
    requested_ranks = params.get("ranks")
    ranks = (
        {int(rank) for rank in requested_ranks}
        if requested_ranks
        else {int(item["rank"]) for item in manifest["candidates"]}
    )
    for rank in ranks:
        candidate_by_rank(manifest, rank)

    hits: list[dict[str, Any]] = []
    lowered_terms = [term.casefold() for term in terms]
    for rank in sorted(ranks):
        candidate = candidate_by_rank(manifest, rank)
        path = local_path(task_dir, candidate["extracted_path"])
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        heading = ""
        for line_number, line in enumerate(lines, start=1):
            stripped = line.strip()
            if stripped.startswith("#"):
                heading = stripped[:240]
            folded = stripped.casefold()
            matched = [
                term for term, lowered in zip(terms, lowered_terms) if lowered in folded
            ]
            if not matched:
                continue
            context_start = max(0, line_number - 2)
            context_end = min(len(lines), line_number + 1)
            snippet = " ".join(
                part.strip() for part in lines[context_start:context_end] if part.strip()
            )
            hits.append(
                {
                    "rank": rank,
                    "document_id": candidate["document_id"],
                    "line": line_number,
                    "heading": heading,
                    "matched_terms": matched,
                    "snippet": snippet[:900],
                }
            )
            if len(hits) >= max_hits:
                break
        if len(hits) >= max_hits:
            break
    return {
        "message": json.dumps(
            {"terms": terms, "hit_count": len(hits), "hits": hits},
            ensure_ascii=False,
            indent=2,
        ),
        "details": {"hit_count": len(hits)},
    }


def set_direction(
    task_dir: Path, manifest: dict[str, Any], params: dict[str, Any]
):
    alternatives = params.get("alternatives", [])
    if not 2 <= len(alternatives) <= 5:
        raise ValueError("provide 2-5 materially different task directions")
    titles = []
    normalized = []
    for alternative in alternatives:
        title = str(alternative.get("title") or "").strip()
        thesis = str(alternative.get("thesis") or "").strip()
        ranks = sorted({int(rank) for rank in alternative.get("candidateRanks", [])})
        risks = [
            str(item).strip()
            for item in alternative.get("risks", [])
            if str(item).strip()
        ]
        if not title or len(thesis) < 80:
            raise ValueError("each direction needs a title and a detailed thesis")
        if not 7 <= len(ranks) <= 17:
            raise ValueError("each direction must identify 7-17 candidate ranks")
        for rank in ranks:
            candidate_by_rank(manifest, rank)
        if not risks:
            raise ValueError("each direction must identify at least one material risk")
        titles.append(title)
        normalized.append(
            {
                "title": title,
                "thesis": thesis,
                "candidate_ranks": ranks,
                "risks": risks,
            }
        )
    if len(titles) != len(set(titles)):
        raise ValueError("direction titles must be unique")

    selected_title = str(params.get("selectedTitle") or "").strip()
    if selected_title not in titles:
        raise ValueError("selectedTitle must match one proposed direction")
    selection_reason = str(params.get("selectionReason") or "").strip()
    if len(selection_reason) < 80:
        raise ValueError("selectionReason must explain evidence and tradeoffs")

    payload = {
        "schema_version": "1.0",
        "task_id": manifest["task_id"],
        "alternatives": normalized,
        "selected_title": selected_title,
        "selection_reason": selection_reason,
        "selected_direction": normalized[titles.index(selected_title)],
    }
    with mutation_lock(task_dir):
        write_json(task_dir / "working" / "direction_plan.json", payload)
    return {
        "message": (
            f"Saved {len(normalized)} directions; selected: {selected_title}. "
            "You may now create auxiliary files or assemble attachments."
        ),
        "details": payload,
    }


def scalar(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if not isinstance(value, str):
        raise ValueError("table cells must be strings, numbers, booleans or null")
    if len(value) > 5000:
        raise ValueError("table cell exceeds 5000 characters")
    if value.startswith("="):
        raise ValueError("generated auxiliary files must not contain formulas")
    return value


def parse_payload(value: str) -> Any:
    try:
        return json.loads(value)
    except json.JSONDecodeError as error:
        raise ValueError(f"payload must be valid JSON for this format: {error}") from error


def write_generated_file(
    path: Path,
    file_format: str,
    payload_text: str,
    purpose: str,
    source_ids: list[str],
) -> None:
    source_label = ", ".join(source_ids) if source_ids else "无；本文件为明确任务假设"
    if file_format in {"markdown", "text"}:
        if len(payload_text.strip()) < 50:
            raise ValueError("generated text content is too short")
        prefix = (
            "文件性质：任务设定/来源整理生成附件，不是外部出版物。\n"
            f"用途：{purpose}\n"
            f"来源候选：{source_label}\n\n"
        )
        path.write_text(prefix + payload_text.strip() + "\n", encoding="utf-8")
        return

    payload = parse_payload(payload_text)
    if file_format == "csv":
        rows = payload.get("rows") if isinstance(payload, dict) else None
        if not isinstance(rows, list) or not rows:
            raise ValueError('CSV payload must be {"rows": [[...], ...]}')
        if len(rows) > 5000:
            raise ValueError("CSV row limit is 5000")
        with path.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(stream)
            writer.writerow(["文件性质", "任务设定/来源整理生成附件，不是外部出版物"])
            writer.writerow(["用途", purpose])
            writer.writerow(["来源候选", source_label])
            writer.writerow([])
            for row in rows:
                if not isinstance(row, list) or len(row) > 100:
                    raise ValueError("each CSV row must be an array of at most 100 cells")
                writer.writerow([scalar(value) for value in row])
        return

    sheets = payload.get("sheets") if isinstance(payload, dict) else None
    if not isinstance(sheets, list) or not 1 <= len(sheets) <= 10:
        raise ValueError('XLSX payload must contain 1-10 entries in "sheets"')
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    info = workbook.create_sheet("说明")
    info_rows = [
        ["文件性质", "任务设定/来源整理生成附件，不是外部出版物"],
        ["用途", purpose],
        ["来源候选", source_label],
    ]
    for row in info_rows:
        info.append(row)
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 80
    for cell in info[1]:
        cell.font = Font(bold=True)

    used_names = {"说明"}
    for sheet_spec in sheets:
        name = str(sheet_spec.get("name") or "").strip()
        rows = sheet_spec.get("rows")
        if (
            not name
            or len(name) > 31
            or re.search(r"[:\\/?*\[\]]", name)
            or name in used_names
        ):
            raise ValueError(f"invalid or duplicate worksheet name: {name!r}")
        if not isinstance(rows, list) or not rows or len(rows) > 5000:
            raise ValueError(f"worksheet {name} must contain 1-5000 rows")
        used_names.add(name)
        worksheet = workbook.create_sheet(name)
        max_columns = 0
        for row in rows:
            if not isinstance(row, list) or len(row) > 100:
                raise ValueError(
                    f"worksheet {name} rows must have at most 100 cells"
                )
            values = [scalar(value) for value in row]
            worksheet.append(values)
            max_columns = max(max_columns, len(values))
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for column in range(1, max_columns + 1):
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(column)
            ].width = 18
    workbook.save(path)
    workbook.close()


def create_generated(
    task_dir: Path, manifest: dict[str, Any], params: dict[str, Any]
):
    if not (task_dir / "working" / "direction_plan.json").is_file():
        raise ValueError("set_task_direction must be completed first")
    filename = clean_filename(params["filename"])
    file_format = str(params["format"])
    expected_extension = GENERATED_EXTENSIONS.get(file_format)
    if expected_extension is None or Path(filename).suffix.lower() != expected_extension:
        raise ValueError(
            f"filename extension must be {expected_extension} for format {file_format}"
        )
    purpose = str(params.get("purpose") or "").strip()
    if len(purpose) < 20:
        raise ValueError("purpose must explain why this generated file is necessary")
    source_ids = sorted(
        {
            str(document_id).strip()
            for document_id in params.get("sourceDocumentIds", [])
            if str(document_id).strip()
        }
    )
    for document_id in source_ids:
        candidate_by_id(manifest, document_id)
    payload_text = str(params.get("payload") or "")

    generated_dir = task_dir / "generated"
    generation_manifest_path = generated_dir / "_generation_manifest.json"
    with mutation_lock(task_dir):
        generated_dir.mkdir(parents=True, exist_ok=True)
        generation_manifest = (
            json.loads(generation_manifest_path.read_text(encoding="utf-8"))
            if generation_manifest_path.is_file()
            else {"schema_version": "1.0", "files": []}
        )
        existing = {
            item["filename"]: item for item in generation_manifest.get("files", [])
        }
        if filename not in existing and len(existing) >= 3:
            raise ValueError("at most 3 generated auxiliary attachments are allowed")
        target = generated_dir / filename
        temporary = generated_dir / f".{filename}.tmp"
        write_generated_file(
            temporary,
            file_format,
            payload_text,
            purpose,
            source_ids,
        )
        temporary.replace(target)
        existing[filename] = {
            "filename": filename,
            "format": file_format,
            "purpose": purpose,
            "source_document_ids": source_ids,
            "sha256": sha256(target),
        }
        generation_manifest["files"] = sorted(
            existing.values(), key=lambda item: item["filename"]
        )
        write_json(generation_manifest_path, generation_manifest)
    return {
        "message": (
            f"Generated {filename} ({target.stat().st_size} bytes). "
            "The file is staged under generated/ and is not final until assembly."
        ),
        "details": existing[filename],
    }


def assemble(
    task_dir: Path, manifest: dict[str, Any], params: dict[str, Any]
):
    direction_path = task_dir / "working" / "direction_plan.json"
    if not direction_path.is_file():
        raise ValueError("set_task_direction must be completed first")
    direction = json.loads(direction_path.read_text(encoding="utf-8"))
    allowed_ranks = set(direction["selected_direction"]["candidate_ranks"])
    specs = params.get("attachments", [])
    if not 10 <= len(specs) <= 17:
        raise ValueError("final attachments must contain 10-17 files")

    generation_manifest_path = task_dir / "generated" / "_generation_manifest.json"
    generation_manifest = (
        json.loads(generation_manifest_path.read_text(encoding="utf-8"))
        if generation_manifest_path.is_file()
        else {"files": []}
    )
    generated_by_name = {
        item["filename"]: item for item in generation_manifest.get("files", [])
    }

    seen_sources: set[str] = set()
    normalized: list[dict[str, Any]] = []
    candidate_ranks: set[int] = set()
    generated_count = 0
    for spec in specs:
        role = str(spec.get("role") or "")
        rationale = str(spec.get("rationale") or "").strip()
        expected_use = str(spec.get("expectedUse") or "").strip()
        if role not in ROLES:
            raise ValueError(f"invalid attachment role: {role}")
        if len(rationale) < 20 or len(expected_use) < 10:
            raise ValueError("each attachment needs rationale and expectedUse")
        rank = spec.get("candidateRank")
        generated_filename = spec.get("generatedFilename")
        if (rank is None) == (generated_filename is None):
            raise ValueError(
                "each attachment must set exactly one of candidateRank or "
                "generatedFilename"
            )
        if rank is not None:
            rank = int(rank)
            if rank not in allowed_ranks:
                raise ValueError(
                    f"candidate rank {rank} is outside the selected direction"
                )
            if rank in candidate_ranks:
                raise ValueError(f"duplicate candidate rank: {rank}")
            if role == "generated":
                raise ValueError("candidate files cannot use the generated role")
            candidate = candidate_by_rank(manifest, rank)
            candidate_ranks.add(rank)
            source_key = f"candidate:{rank}"
            normalized.append(
                {
                    "source_key": source_key,
                    "origin": "candidate",
                    "rank": rank,
                    "source_path": local_path(task_dir, candidate["candidate_path"]),
                    "filename": Path(candidate["candidate_path"]).name,
                    "document_id": candidate["document_id"],
                    "role": role,
                    "rationale": rationale,
                    "expected_use": expected_use,
                }
            )
        else:
            filename = clean_filename(str(generated_filename))
            if filename not in generated_by_name:
                raise ValueError(f"generated file was not staged: {filename}")
            if role != "generated":
                raise ValueError("generated files must use the generated role")
            generated_count += 1
            source_key = f"generated:{filename}"
            generated = generated_by_name[filename]
            normalized.append(
                {
                    "source_key": source_key,
                    "origin": "generated",
                    "source_path": task_dir / "generated" / filename,
                    "filename": filename,
                    "role": "generated",
                    "rationale": rationale,
                    "expected_use": expected_use,
                    "purpose": generated["purpose"],
                    "source_document_ids": generated["source_document_ids"],
                }
            )
        if source_key in seen_sources:
            raise ValueError(f"duplicate attachment source: {source_key}")
        seen_sources.add(source_key)
    if generated_count > 3:
        raise ValueError("at most 3 generated attachments are allowed")
    if sum(item["role"] == "core" for item in normalized) < 2:
        raise ValueError("the attachment set needs at least 2 core evidence files")
    final_dir = task_dir / "final"
    with mutation_lock(task_dir):
        staging = Path(
            tempfile.mkdtemp(prefix=".final-staging-", dir=str(task_dir))
        )
        try:
            attachments_dir = staging / "attachments"
            internal_dir = staging / "internal"
            attachments_dir.mkdir()
            internal_dir.mkdir()
            selection_records = []
            final_filenames = []
            for position, item in enumerate(normalized, start=1):
                source_path = item["source_path"]
                if not source_path.is_file():
                    raise FileNotFoundError(source_path)
                final_filename = final_attachment_filename(
                    position,
                    item["filename"],
                )
                final_filenames.append(final_filename)
                target = attachments_dir / final_filename
                shutil.copy2(source_path, target)
                record = {
                    key: value
                    for key, value in item.items()
                    if key not in {"source_key", "rank", "source_path", "filename"}
                }
                record["source_filename"] = item["filename"]
                record["filename"] = final_filename
                record["sha256"] = sha256(target)
                selection_records.append(record)

            excluded = [
                {
                    "rank": int(candidate["rank"]),
                    "document_id": candidate["document_id"],
                    "filename": candidate["attachment_filename"],
                    "reason": "未纳入已选题目方向的最终附件组合",
                }
                for candidate in manifest["candidates"]
                if int(candidate["rank"]) not in candidate_ranks
            ]
            selection = {
                "schema_version": "1.0",
                "task_id": manifest["task_id"],
                "direction": direction["selected_title"],
                "total_attachments": len(selection_records),
                "original_count": len(selection_records) - generated_count,
                "generated_count": generated_count,
                "attachments": selection_records,
                "excluded_candidates": excluded,
            }
            write_json(internal_dir / "selection_manifest.json", selection)
            shutil.copy2(direction_path, internal_dir / "direction_plan.json")
            if final_dir.exists():
                shutil.rmtree(final_dir)
            staging.replace(final_dir)
        except Exception:
            if staging.exists():
                shutil.rmtree(staging)
            raise
    return {
        "message": (
            f"Assembled {len(normalized)} final attachments "
            f"({len(normalized) - generated_count} candidate, "
            f"{generated_count} generated). Now derive and finalize the "
            "workflow and query."
        ),
        "details": {
            "attachment_count": len(normalized),
            "generated_count": generated_count,
            "filenames": final_filenames,
        },
    }


def finalize(task_dir: Path, manifest: dict[str, Any], params: dict[str, Any]):
    final_dir = task_dir / "final"
    selection_path = final_dir / "internal" / "selection_manifest.json"
    if not selection_path.is_file():
        raise ValueError("assemble_final_attachments must be completed first")
    query = str(params.get("queryMarkdown") or "").strip()
    workflow = str(params.get("workflowMarkdown") or "").strip()
    evidence = str(params.get("evidenceMatrixMarkdown") or "").strip()
    quality = str(params.get("qualityReviewMarkdown") or "").strip()
    if len(query) < 500:
        raise ValueError("queryMarkdown is too short for a complex professional task")
    if len(workflow) < 400:
        raise ValueError("workflowMarkdown is too short for a 12-step workflow")
    if len(evidence) < 300:
        raise ValueError("evidenceMatrixMarkdown is too short")
    if len(quality) < 300:
        raise ValueError("qualityReviewMarkdown is too short")

    with mutation_lock(task_dir):
        write_json(final_dir / "query.json", [{"query": query}])
        (final_dir / "query.md").write_text(query + "\n", encoding="utf-8")
        (final_dir / "workflow.md").write_text(
            workflow + "\n",
            encoding="utf-8",
        )
        (final_dir / "internal" / "evidence_matrix.md").write_text(
            evidence + "\n", encoding="utf-8"
        )
        (final_dir / "internal" / "quality_review.md").write_text(
            quality + "\n", encoding="utf-8"
        )
        result = validate_task(task_dir)
    return {
        "message": (
            "Stage 2 task package passed deterministic validation: "
            + json.dumps(result, ensure_ascii=False)
        ),
        "details": result,
        "terminate": True,
    }


ACTIONS = {
    "candidate_inventory": inventory,
    "read_candidate": read_candidate,
    "search_evidence": search_evidence,
    "set_task_direction": set_direction,
    "create_generated_attachment": create_generated,
    "assemble_final_attachments": assemble,
    "finalize_task": finalize,
}


def handle(request: dict[str, Any]) -> dict[str, Any]:
    action = request.get("action")
    if action not in ACTIONS:
        raise ValueError(f"unknown Pi tool action: {action}")
    task_dir, manifest = task_context(request["cwd"])
    params = request.get("params") or {}
    return ACTIONS[action](task_dir, manifest, params)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--request-json", required=True)
    args = parser.parse_args()
    try:
        request = json.loads(args.request_json)
        response = handle(request)
    except Exception as error:
        print(
            json.dumps(
                {"ok": False, "error": f"{type(error).__name__}: {error}"},
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        raise SystemExit(1) from error
    print(json.dumps({"ok": True, **response}, ensure_ascii=False))


if __name__ == "__main__":
    main()
